import os
import xml.etree.ElementTree as ET
import datetime
from dataclasses import dataclass
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path

@dataclass
class Fattura:
    cedente_id_fiscale: str
    cedente_denominazione: str
    cedente_regime_fiscale: str
    divisa: str
    data: datetime.date
    numero: str
    data_scadenza_pagamento: datetime.date
    importo_pagamento: float

@dataclass
class TotaleFattureCedente:
    cedente_id_fiscale: str
    cedente_denominazione: str
    totale_pagamenti: float

def parse_date(date_str: str) -> datetime.date:
    return datetime.datetime.strptime(date_str, '%Y-%m-%d').date()

def process_xml_file(file_path: str) -> Fattura:
    """Process a single XML file and return a Fattura object."""
    tree = ET.parse(file_path)
    root = tree.getroot()
    
    try:
        # Find header
        header = root.find('FatturaElettronicaHeader')
        if header is None:
            raise ValueError("FatturaElettronicaHeader not found")
            
        # Find CedentePrestatore using relative path from header
        cedente = header.find('CedentePrestatore')
        if cedente is None:
            raise ValueError("CedentePrestatore not found")
            
        # Extract Cedente information
        dati_anagrafici = cedente.find('DatiAnagrafici')
        if dati_anagrafici is None:
            raise ValueError("DatiAnagrafici not found")
        
        # Find IdFiscaleIVA
        id_fiscale_iva = dati_anagrafici.find('IdFiscaleIVA/IdCodice')
        if id_fiscale_iva is None:
            raise ValueError("IdFiscaleIVA/IdCodice not found")
        id_fiscale = id_fiscale_iva.text
        
        # Find Denominazione
        denominazione = dati_anagrafici.find('Anagrafica/Denominazione')
        if denominazione is None:
            raise ValueError("Anagrafica/Denominazione not found")
        denominazione = denominazione.text
        
        # Find RegimeFiscale
        regime_fiscale = dati_anagrafici.find('RegimeFiscale')
        if regime_fiscale is None:
            raise ValueError("RegimeFiscale not found")
        regime_fiscale = regime_fiscale.text
        
        # Extract FatturaElettronicaBody information
        body = root.find('FatturaElettronicaBody')
        if body is None:
            raise ValueError("FatturaElettronicaBody not found")
            
        # Extract DatiGenerali
        dati_generali = body.find('DatiGenerali/DatiGeneraliDocumento')
        if dati_generali is None:
            raise ValueError("DatiGeneraliDocumento not found")
        
        divisa = dati_generali.find('Divisa').text
        data = parse_date(dati_generali.find('Data').text)
        numero = dati_generali.find('Numero').text
        
        # Extract DatiPagamento
        dati_pagamento = body.find('DatiPagamento/DettaglioPagamento')
        if dati_pagamento is None:
            raise ValueError("DatiPagamento/DettaglioPagamento not found")
        
        data_scadenza = parse_date(dati_pagamento.find('DataScadenzaPagamento').text)
        importo = float(dati_pagamento.find('ImportoPagamento').text)
        
        return Fattura(
            cedente_id_fiscale=id_fiscale,
            cedente_denominazione=denominazione,
            cedente_regime_fiscale=regime_fiscale,
            divisa=divisa,
            data=data,
            numero=numero,
            data_scadenza_pagamento=data_scadenza,
            importo_pagamento=importo
        )
        
    except Exception as e:
        print(f"Detailed error in {file_path}: {str(e)}")
        raise

def process_folder(folder_path: str, start_date: datetime.date, end_date: datetime.date) -> List[Fattura]:
    """Process all XML files in the folder and subfolders."""
    fatture = []
    
    for root, _, files in os.walk(folder_path):
        for file in files:
            if file.endswith('.xml'):
                file_path = os.path.join(root, file)
                try:
                    # Parse and process the file
                    fattura = process_xml_file(file_path)
                    
                    # Check if the invoice date is within the specified range
                    if start_date <= fattura.data <= end_date:
                        fatture.append(fattura)
                        print(f"Successfully processed: {file_path}")
                        
                except (ET.ParseError, AttributeError, ValueError) as e:
                    print(f"Error processing file {file_path}: {str(e)}")
                    continue
    
    return fatture

def aggregate_by_cedente(fatture: List[Fattura]) -> List[TotaleFattureCedente]:
    """Aggregate fatture by cedente and calculate totals."""
    totals = {}
    
    for fattura in fatture:
        key = (fattura.cedente_id_fiscale, fattura.cedente_denominazione)
        if key not in totals:
            totals[key] = 0
        totals[key] += fattura.importo_pagamento
    
    return [
        TotaleFattureCedente(
            cedente_id_fiscale=id_fiscale,
            cedente_denominazione=denominazione,
            totale_pagamenti=total
        )
        for (id_fiscale, denominazione), total in totals.items()
    ]

def write_excel(totali: List[TotaleFattureCedente], fatture: List[Fattura], output_file: str):
    """Write the data to an Excel file with two sheets: summary and details."""
    wb = Workbook()
    
    # Create Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Riepilogo per Fornitore"
    
    # Add headers with styling
    headers_summary = ['Cedente.IdFiscaleIVA', 'Cedente.Anagrafica.Denominazione', 'Totale Pagamenti']
    for col, header in enumerate(headers_summary, 1):
        cell = ws_summary.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Add data to summary sheet
    for row, totale in enumerate(totali, 2):
        ws_summary.cell(row=row, column=1, value=totale.cedente_id_fiscale)
        ws_summary.cell(row=row, column=2, value=totale.cedente_denominazione)
        cell = ws_summary.cell(row=row, column=3, value=totale.totale_pagamenti)
        cell.number_format = '#,##0.00 €'
    
    # Create Details sheet
    ws_details = wb.create_sheet("Dettaglio Fatture")
    
    # Add headers with styling
    headers_details = [
        'Cedente.IdFiscaleIVA',
        'Cedente.Denominazione',
        'Regime Fiscale',
        'Divisa',
        'Data Fattura',
        'Numero Fattura',
        'Data Scadenza',
        'Importo'
    ]
    
    for col, header in enumerate(headers_details, 1):
        cell = ws_details.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Add data to details sheet
    for row, fattura in enumerate(fatture, 2):
        ws_details.cell(row=row, column=1, value=fattura.cedente_id_fiscale)
        ws_details.cell(row=row, column=2, value=fattura.cedente_denominazione)
        ws_details.cell(row=row, column=3, value=fattura.cedente_regime_fiscale)
        ws_details.cell(row=row, column=4, value=fattura.divisa)
        ws_details.cell(row=row, column=5, value=fattura.data)
        ws_details.cell(row=row, column=6, value=fattura.numero)
        ws_details.cell(row=row, column=7, value=fattura.data_scadenza_pagamento)
        cell = ws_details.cell(row=row, column=8, value=fattura.importo_pagamento)
        cell.number_format = '#,##0.00 €'
    
    # Adjust column widths
    for ws in [ws_summary, ws_details]:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save the workbook
    wb.save(output_file)

def main(folder_path: str, start_date_str: str, end_date_str: str):
    """Main function to process invoices and generate Excel file."""
    # Convert date strings to date objects
    start_date = parse_date(start_date_str)
    end_date = parse_date(end_date_str)
    
    print(f"Processing files from {start_date} to {end_date}")
    print(f"Looking in folder: {folder_path}")
    
    # Process all files
    fatture = process_folder(folder_path, start_date, end_date)
    
    # Aggregate by cedente
    totali = aggregate_by_cedente(fatture)
    
    # Write output Excel
    output_file = "fattura-pa-summary.xlsx"
    write_excel(totali, fatture, output_file)
    print(f"\nProcessed {len(fatture)} invoices")
    print(f"Generated summary for {len(totali)} suppliers in {output_file}")

if __name__ == "__main__":
    import sys
    if len(sys.argv) != 4:
        print("Usage: python script.py <folder_path> <start_date> <end_date>")
        print("Dates should be in YYYY-MM-DD format")
        sys.exit(1)
        
    main(sys.argv[1], sys.argv[2], sys.argv[3])